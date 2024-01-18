import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def extrair_informacoes_do_produto(produto):
    nome_produto = produto.find('h2', class_='Text_Text__ARJdp Text_MobileLabelXs__dHwGG Text_DesktopLabelSAtLarge__wWsED ProductCard_ProductCard_Name__U_mUQ').text
    preco_produto_texto = produto.find('p', class_='Text_Text__ARJdp Text_MobileHeadingS__HEz7L').text
    preco_produto = float(preco_produto_texto.replace('R$', '').replace('.', '').replace(',', '.'))
    link_produto = 'https://www.zoom.com.br' + produto.find('a', class_='ProductCard_ProductCard_Inner__gapsh')['href']

    return nome_produto, preco_produto, link_produto

def obter_informacoes_do_site(url):
    response = requests.get(url)

    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        produtos = soup.find_all('div', class_='Paper_Paper__4XALQ Paper_Paper__bordered__cl5Rh Card_Card__Zd8Ef Card_Card__clicable__ewI68 ProductCard_ProductCard__WWKKW')
        informacoes_produtos = [extrair_informacoes_do_produto(produto) for produto in produtos]

        return informacoes_produtos

def salva_em_excel(informacoes_produtos, nome_arquivo):
    workbook = Workbook()
    sheet = workbook.active
    
    cabecalhos = ['Nome do Produto', 'Preço', 'Link do produto']
    for col_num, cabecalho in enumerate(cabecalhos, start=1):
        sheet[get_column_letter(col_num) + '1'] = cabecalho

    for row_num, informacoes_produto in enumerate(informacoes_produtos, start=sheet.max_row + 1):
        for col_num, valor in enumerate(informacoes_produto, start=1):
            sheet[get_column_letter(col_num) + str(row_num)] = valor
             
    workbook.save(nome_arquivo)

def pesquisa_produto():
    produto = input('Digite o produto que deseja pesquisar: ')
    return 'https://www.zoom.com.br/search?q=' + produto

if __name__ == "__main__":
    nome_do_arquivo_excel = 'resultado.xlsx'
    url_do_site = pesquisa_produto()

    informacoes_produtos = obter_informacoes_do_site(url_do_site)
    
    if informacoes_produtos:
        # Ordena as informações dos produtos com base nos preços
        informacoes_produtos_ordenadas = sorted(informacoes_produtos, key=lambda x: x[1])
        
        # Seleciona os 5 menores preços
        cinco_menores_precos = informacoes_produtos_ordenadas[:5]
        
        # Salva as informações dos 5 menores preços em um arquivo Excel
        salva_em_excel(cinco_menores_precos, nome_do_arquivo_excel)
        
        print(f"As informações dos 5 produtos com os menores preços foram salvas em '{nome_do_arquivo_excel}'.")